#!/usr/bin/env python3
"""
State Budget Bill Analyzer
Extracts agency/department appropriations from a PDF using OpenAI GPT-4o,
then produces a formatted Excel summary report.

Requires:
  - OPENAI_API_KEY environment variable
  - pip install openai pdfplumber openpyxl
"""

import sys
import os
import json
import re
from collections import deque
from pathlib import Path
from datetime import datetime
import pdfplumber
import openai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Extraction (GPT-4o)
# ---------------------------------------------------------------------------

def extract_budget_data(pdf_path: str) -> dict:
    warnings = []

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("Error: OPENAI_API_KEY is not set.")
        print("Add it to ~/.zshrc:  export OPENAI_API_KEY=\"sk-...\"")
        sys.exit(1)

    client = openai.OpenAI(api_key=api_key)

    # Build page-aware chunks (~40K chars each).
    # Chunks are stored as lists of (page_num, tagged_text) so they can be
    # split in half and retried automatically when GPT-4o hits the token limit.
    PAGE_CHUNK_LIMIT = 40_000
    initial_chunks = []   # list of [(page_num, tagged_text), ...]
    current_pages  = []
    current_len    = 0
    page_count     = 0

    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)
        for page_num, page in enumerate(pdf.pages, 1):
            page_text = page.extract_text() or ""
            if not page_text.strip():
                warnings.append(f"Page {page_num}: no extractable text (may be scanned).")
                continue
            tagged = f"--- PAGE {page_num} ---\n{page_text}"
            tagged_len = len(tagged)
            if current_pages and current_len + tagged_len > PAGE_CHUNK_LIMIT:
                initial_chunks.append(current_pages)
                current_pages = []
                current_len   = 0
            current_pages.append((page_num, tagged))
            current_len += tagged_len
        if current_pages:
            initial_chunks.append(current_pages)

    system_prompt = (
        "You are a state budget bill analyst. Extract every named appropriation from the "
        "provided state budget bill text.\n\n"
        "Return ONLY valid JSON in exactly this shape:\n"
        "{\n"
        '  "fiscal_years": ["2026-27"],\n'
        '  "section_totals": {\n'
        '    "Department Of Agriculture – General Administration": {"2026-27": 14719072.0},\n'
        '    "Department Of Agriculture – Meat Inspection Fund": {"2026-27": 1214444.0},\n'
        '    "State Board Of Education": {"2026-27": 3100000000.0}\n'
        "  },\n"
        '  "sub_allocations": {\n'
        '    "Alabama Reading Initiative": {"2026-27": 151000000.0},\n'
        '    "Alabama Numeracy Act": {"2026-27": 114000000.0}\n'
        "  }\n"
        "}\n\n"
        "TWO BILL FORMATS — handle both:\n\n"
        "FORMAT A — TEXT/NARRATIVE (common in Oklahoma and many other states):\n"
        "  Each appropriation is stated in a numbered section of prose, e.g.:\n"
        "  'SECTION 1. There is hereby appropriated to the State Board of Education "
        "the sum of $1,704,307,500'\n"
        "  'SECTION 2. There is hereby appropriated to the Department of Transportation "
        "$610,000,000'\n"
        "  Every such statement is an independent appropriation. Put each named recipient "
        "and its stated amount in section_totals. Do NOT skip any section — even if the "
        "sections run consecutively with no blank lines between them.\n\n"
        "FORMAT B — TABLE/STRUCTURED (common in West Virginia and similar states):\n"
        "  Each agency has a titled section block with line items and a 'Total' row. "
        "  Named fund divisions separated by a dash (e.g. '– Meat Inspection Fund', "
        "'– Custodial Fund') are INDEPENDENT sections with their own fund code and Total "
        "— put every one in section_totals.\n\n"
        "SECTION TOTALS — put in 'section_totals':\n"
        "Any named government entity or fund that receives a direct, independent "
        "appropriation — regardless of whether the bill states it as a table row, "
        "a prose sentence, or a numbered bill section. Include:\n"
        "- Government departments, agencies, boards, commissions, bureaus\n"
        "- Universities, colleges, and educational authorities\n"
        "- Named funds or authorities with their own appropriated amount\n"
        "- Dash-separated fund divisions in table-format bills\n"
        "- Law enforcement, public safety, corrections entities and their funds\n"
        "- Any other named entity with its own stated appropriation\n\n"
        "SEPARATE SECTIONS RULE — never merge across sections:\n"
        "If the same entity name appears in multiple independent bill sections (e.g. a General\n"
        "Revenue article AND a Special Revenue article), return each as its own entry with its\n"
        "exact stated amount. Never add or merge figures from separate sections.\n\n"
        "SUB-ALLOCATIONS — put in 'sub_allocations':\n"
        "Named line items that appear WITHIN another entity's appropriation block and show "
        "how that entity's total is divided among programs or purposes. These are spending "
        "directions inside an already-counted total — do NOT put them in section_totals.\n"
        "Example: 'Alabama Reading Initiative — $151M' and 'Alabama Numeracy Act — $114M'\n"
        "listed inside 'State Board Of Education — Total $3.1B' → sub_allocations.\n"
        "If the bill only states a lump-sum appropriation for an entity with no internal "
        "breakdown, sub_allocations will be empty — that is fine.\n\n"
        "EXCLUDE from both dicts:\n"
        "- Individual persons' names (e.g. 'Smith, John')\n"
        "- Grand-total rollup lines that sum multiple agencies "
        "(e.g. 'Total General Fund', 'All Funds Total', 'Grand Total')\n"
        "- Private nonprofit organizations that receive grants routed THROUGH a department\n\n"
        "Other rules:\n"
        "- fiscal_years: fiscal-year labels in this chunk, formatted YYYY-YY (e.g. 2026-27).\n"
        "- Entity names in Title Case.\n"
        "- Dollar amounts as plain floats — no $ signs, no commas. "
        "If amounts are in thousands, multiply by 1000.\n"
        "- If no appropriations are found, return "
        '{"fiscal_years": [], "section_totals": {}, "sub_allocations": {}}.'
    )

    # Separate tracking for section_totals and sub_allocations.
    # Each category has its own canonical-name map, entities dict, pages dict,
    # and exact-dedup set so the two pools never interfere.
    st_canonical: dict = {}   # section_totals
    st_entities:  dict = {}
    st_pages:     dict = {}
    st_seen:      set  = set()

    sa_canonical: dict = {}   # sub_allocations
    sa_entities:  dict = {}
    sa_pages:     dict = {}
    sa_seen:      set  = set()

    all_fiscal_years: list = []

    work_queue = deque(initial_chunks)
    print(f"  Extracting with GPT-4o — {len(initial_chunks)} chunk(s)...")

    while work_queue:
        pages = work_queue.popleft()
        start_page = pages[0][0]
        end_page   = pages[-1][0]
        label      = f"pages {start_page}–{end_page} of {page_count}"
        chunk_text = '\n'.join(text for _, text in pages)

        print(f"    Chunk: {label}")
        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                max_tokens=16384,
                temperature=0,
                seed=42,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Extract appropriations ({label}):\n\n{chunk_text}"},
                ],
            )
            finish_reason = response.choices[0].finish_reason
            raw = response.choices[0].message.content or ""

            if finish_reason == "length":
                if len(pages) > 1:
                    mid = len(pages) // 2
                    work_queue.appendleft(pages[mid:])
                    work_queue.appendleft(pages[:mid])
                    print(f"      → token limit hit — splitting into {mid} + {len(pages) - mid} pages")
                    continue
                else:
                    warnings.append(
                        f"{label}: single-page chunk hit token limit — entities on this page may be incomplete."
                    )
                    print(f"      WARNING: single-page chunk hit token limit")

            data = json.loads(raw)

            for fy in data.get("fiscal_years", []):
                if fy not in all_fiscal_years:
                    all_fiscal_years.append(fy)

            # Process both categories with identical dedup logic
            categories = [
                (data.get("section_totals", {}), st_canonical, st_entities, st_pages, st_seen),
                (data.get("sub_allocations", {}),  sa_canonical, sa_entities, sa_pages, sa_seen),
            ]
            chunk_st_count = 0
            chunk_sa_count = 0

            for cat_idx, (src_dict, canonical, entities, pg_map, seen) in enumerate(categories):
                for name, amounts in src_dict.items():
                    if not isinstance(amounts, dict):
                        continue

                    key = name.strip().lower()
                    valid_amounts = {}
                    for fy, amt in amounts.items():
                        try:
                            valid_amounts[fy] = float(amt)
                        except (TypeError, ValueError):
                            continue
                    if not valid_amounts or not any(v > 0 for v in valid_amounts.values()):
                        continue

                    amounts_sig = frozenset(valid_amounts.items())
                    if (key, amounts_sig) in seen:
                        continue
                    seen.add((key, amounts_sig))

                    if key not in canonical:
                        canonical[key] = name.strip()
                        entities[key]  = valid_amounts
                        pg_map[key]    = start_page
                    else:
                        # Same name, different amounts → separate bill section
                        page_key = f"{key} [p{start_page}]"
                        counter = 2
                        while page_key in canonical:
                            page_key = f"{key} [p{start_page}:{counter}]"
                            counter += 1
                        canonical[page_key] = f"{name.strip()} (p. {start_page})"
                        entities[page_key]  = valid_amounts
                        pg_map[page_key]    = start_page

                    if cat_idx == 0:
                        chunk_st_count += 1
                    else:
                        chunk_sa_count += 1

            print(f"      → {chunk_st_count} section totals, {chunk_sa_count} sub-allocations found")
            if chunk_st_count == 0 and chunk_sa_count == 0:
                warnings.append(
                    f"{label}: 0 entities returned — this page range may be missing from the report."
                )
                print("      WARNING: no entities found in this chunk")

        except Exception as exc:
            warnings.append(f"{label} error: {exc}")
            print(f"    ERROR: {exc}")

    fiscal_years = all_fiscal_years or ['Total']

    # ---------------------------------------------------------------------------
    # Post-processing filters
    # ---------------------------------------------------------------------------

    _personal_name_re = re.compile(r'^[A-Z][A-Za-z\-]+,\s+[A-Z][a-z]+$')
    _page_suffix_re   = re.compile(r'\s*\(p\.\s*[\d:]+\)\s*$')

    def _base_name(name: str) -> str:
        return _page_suffix_re.sub('', name).strip()

    # --- Section-total filters (full suite) ---

    st_display = {k: st_canonical[k] for k in st_entities}

    def _word_set(name: str) -> frozenset:
        return frozenset(re.sub(r'[,\-/]', ' ', _base_name(name).lower()).split())

    word_set_groups: dict = {}
    for key in st_entities:
        ws = _word_set(st_canonical[key])
        word_set_groups.setdefault(ws, []).append(key)

    word_set_dupes: set = set()
    for keys_in_group in word_set_groups.values():
        if len(keys_in_group) > 1:
            base_keys = [k for k in keys_in_group if '[p' not in k]
            if len(base_keys) > 1:
                best = max(base_keys, key=lambda k: sum(st_entities[k].values()))
                for k in base_keys:
                    if k != best:
                        word_set_dupes.add(k)

    def _is_fragment_duplicate(key: str) -> bool:
        name = _base_name(st_display[key]).lower()
        amt  = sum(st_entities[key].values())
        for other_key, other_name in st_display.items():
            if other_key == key:
                continue
            other_base = _base_name(other_name).lower()
            if other_base == name:
                continue
            if name in other_base and other_base != name:
                other_amt = sum(st_entities[other_key].values())
                if other_amt > amt * 10:
                    return True
        return False

    def _should_exclude_st(key: str) -> bool:
        name = st_canonical[key]
        amts = st_entities[key]
        if not any(v > 0 for v in amts.values()):
            return True
        if _personal_name_re.match(name):
            return True
        if key in word_set_dupes:
            return True
        if _is_fragment_duplicate(key):
            return True
        return False

    # --- Sub-allocation filters (personal-name only) ---

    def _should_exclude_sa(key: str) -> bool:
        name = sa_canonical[key]
        amts = sa_entities[key]
        if not any(v > 0 for v in amts.values()):
            return True
        if _personal_name_re.match(name):
            return True
        return False

    ordered_st = {
        st_canonical[k]: v
        for k, v in st_entities.items()
        if not _should_exclude_st(k)
    }

    ordered_sa = {
        sa_canonical[k]: v
        for k, v in sa_entities.items()
        if not _should_exclude_sa(k)
    }

    grand_totals = {
        fy: sum(e.get(fy, 0) for e in ordered_st.values())
        for fy in fiscal_years
    }

    return {
        "section_totals":    ordered_st,
        "sub_allocations":   ordered_sa,
        "fiscal_years":      fiscal_years,
        "grand_totals":      grand_totals,
        "bill_grand_totals": {},
        "page_count":        page_count,
        "warnings":          warnings,
    }


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

COLOR_HEADER_BG  = "1F3864"
COLOR_HEADER_FG  = "FFFFFF"
COLOR_TOTAL_BG   = "D9E1F2"
COLOR_META_FG    = "666666"
COLOR_ROW_EVEN   = "F5F7FA"
COLOR_ROW_ODD    = "FFFFFF"
COLOR_NOTES_FG   = "888888"
COLOR_SA_BG      = "F0F0F0"   # sub-allocation rows — light gray
COLOR_SA_HEADER  = "E0E0E0"   # sub-allocation section header
COLOR_SA_FG      = "555555"

DOLLAR_FMT = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_excel_report(data: dict, source_file: str, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appropriations Summary"

    fys          = data["fiscal_years"]
    entities     = data["section_totals"]
    sub_allocs   = data.get("sub_allocations", {})
    grand_totals = data["grand_totals"]
    bill_totals  = data["bill_grand_totals"]

    FY_START = 2
    num_cols = 1 + len(fys)

    # --- Row 1: Title ---------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value="State Budget Bill — Appropriations Summary")
    c.font      = Font(name="Arial", size=16, bold=True, color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Row 2: Metadata ------------------------------------------------------
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    meta = (f"Source: {Path(source_file).name}   |   "
            f"Pages: {data['page_count']}   |   "
            f"Figures: all-funds TOTAL per entity")
    c = ws.cell(row=2, column=1, value=meta)
    c.font      = Font(name="Arial", size=10, italic=True, color=COLOR_META_FG)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # spacer

    # --- Row 4: Column headers ------------------------------------------------
    ws.row_dimensions[4].height = 22

    c = ws.cell(row=4, column=1, value="Agency / Cabinet / Department")
    c.font      = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_FG)
    c.fill      = _fill(COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _border()

    for col, fy in enumerate(fys, start=FY_START):
        c = ws.cell(row=4, column=col, value=fy)
        c.font      = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_FG)
        c.fill      = _fill(COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = _border()

    # --- Rows 5+: Section-total entity rows -----------------------------------
    for offset, (name, amounts) in enumerate(entities.items()):
        r    = 5 + offset
        fill = COLOR_ROW_EVEN if offset % 2 == 0 else COLOR_ROW_ODD

        c = ws.cell(row=r, column=1, value=name)
        c.font      = Font(name="Arial", size=10)
        c.fill      = _fill(fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()

        for col, fy in enumerate(fys, start=FY_START):
            amt = amounts.get(fy, 0)
            c = ws.cell(row=r, column=col, value=amt if amt else None)
            c.font          = Font(name="Arial", size=10)
            c.fill          = _fill(fill)
            c.alignment     = Alignment(horizontal="right", vertical="center")
            c.border        = _border()
            c.number_format = DOLLAR_FMT

    next_row = 5 + len(entities)

    # --- Named-entity totals row ----------------------------------------------
    c = ws.cell(row=next_row, column=1, value="Total — Named Entity Appropriations")
    c.font      = Font(name="Arial", size=11, bold=True)
    c.fill      = _fill(COLOR_TOTAL_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _border()

    for col, fy in enumerate(fys, start=FY_START):
        amt = grand_totals.get(fy, 0)
        c = ws.cell(row=next_row, column=col, value=amt if amt else None)
        c.font          = Font(name="Arial", size=11, bold=True)
        c.fill          = _fill(COLOR_TOTAL_BG)
        c.alignment     = Alignment(horizontal="right", vertical="center")
        c.border        = _border()
        c.number_format = DOLLAR_FMT

    next_row += 1

    # --- Bill's stated all-funds total (if present) ---------------------------
    if bill_totals:
        c = ws.cell(row=next_row, column=1,
                    value="Bill's Stated All-Funds Total (incl. bonds, transfers, other funds)")
        c.font      = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_FG)
        c.fill      = _fill(COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()

        for col, fy in enumerate(fys, start=FY_START):
            amt = bill_totals.get(fy, 0)
            c = ws.cell(row=next_row, column=col, value=amt if amt else None)
            c.font          = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_FG)
            c.fill          = _fill(COLOR_HEADER_BG)
            c.alignment     = Alignment(horizontal="right", vertical="center")
            c.border        = _border()
            c.number_format = DOLLAR_FMT

        next_row += 1

    # --- Sub-allocations section (shown for reference) ------------------------
    if sub_allocs:
        next_row += 1  # spacer

        ws.merge_cells(start_row=next_row, start_column=1,
                       end_row=next_row, end_column=num_cols)
        c = ws.cell(row=next_row, column=1,
                    value="Sub-Allocations — shown for reference only; already included in section totals above")
        c.font      = Font(name="Arial", size=10, bold=True, italic=True, color=COLOR_SA_FG)
        c.fill      = _fill(COLOR_SA_HEADER)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()
        ws.row_dimensions[next_row].height = 18
        next_row += 1

        for name, amounts in sub_allocs.items():
            c = ws.cell(row=next_row, column=1, value=f"  ↳ {name}")
            c.font      = Font(name="Arial", size=9, italic=True, color=COLOR_SA_FG)
            c.fill      = _fill(COLOR_SA_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = _border()

            for col, fy in enumerate(fys, start=FY_START):
                amt = amounts.get(fy, 0)
                c = ws.cell(row=next_row, column=col, value=amt if amt else None)
                c.font          = Font(name="Arial", size=9, italic=True, color=COLOR_SA_FG)
                c.fill          = _fill(COLOR_SA_BG)
                c.alignment     = Alignment(horizontal="right", vertical="center")
                c.border        = _border()
                c.number_format = DOLLAR_FMT

            next_row += 1

    # --- Warnings -------------------------------------------------------------
    if data["warnings"]:
        next_row += 1
        ws.merge_cells(start_row=next_row, start_column=1,
                       end_row=next_row, end_column=num_cols)
        c = ws.cell(row=next_row, column=1, value="Notes")
        c.font = Font(name="Arial", size=10, bold=True)
        next_row += 1
        for w in data["warnings"]:
            ws.merge_cells(start_row=next_row, start_column=1,
                           end_row=next_row, end_column=num_cols)
            c = ws.cell(row=next_row, column=1, value=f"• {w}")
            c.font      = Font(name="Arial", size=9, color=COLOR_NOTES_FG)
            c.alignment = Alignment(wrap_text=True)
            next_row += 1

    # --- Methodology note -----------------------------------------------------
    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1,
                   end_row=next_row, end_column=num_cols)
    c = ws.cell(row=next_row, column=1,
                value=(
                    "Methodology: Appropriations were extracted from the source PDF using AI (GPT-4o). "
                    "Section totals represent each named agency, department, fund, or authority at its "
                    "own appropriated Total. Sub-allocations (shown below the grand total) are internal "
                    "program lines within a section and are already included in the section total above — "
                    "they are not added to the grand total. Verify all figures against the enrolled bill before citing."
                ))
    c.font      = Font(name="Arial", size=8, italic=True, color=COLOR_NOTES_FG)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[next_row].height = 54

    # --- Column widths and freeze ---------------------------------------------
    ws.column_dimensions["A"].width = 60
    for col in range(FY_START, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "A5"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python budget_parser.py <budget.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = sys.argv[1]

    if not Path(pdf_path).exists():
        print(f"Error: file not found — {pdf_path}")
        sys.exit(1)

    reports_dir = Path(__file__).parent / "Reports"
    reports_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = (sys.argv[2] if len(sys.argv) > 2
                else str(reports_dir / (Path(pdf_path).stem + f"_summary_{timestamp}.xlsx")))

    cache_path = reports_dir / (Path(pdf_path).stem + "_extraction.json")

    print(f"Analyzing: {pdf_path}")
    if cache_path.exists():
        print(f"  Using cached extraction — delete {cache_path.name} to re-run AI extraction")
        with open(cache_path) as f:
            data = json.load(f)
        # Migrate old-format cache (single 'entities' dict) to new two-dict schema
        if "entities" in data and "section_totals" not in data:
            print("  (migrating old cache format to section_totals/sub_allocations)")
            data["section_totals"]  = data.pop("entities")
            data["sub_allocations"] = {}
            data.pop("entity_pages", None)
            # Recompute grand_totals from section_totals to be safe
            data["grand_totals"] = {
                fy: sum(e.get(fy, 0) for e in data["section_totals"].values())
                for fy in data.get("fiscal_years", [])
            }
    else:
        data = extract_budget_data(pdf_path)
        with open(cache_path, 'w') as f:
            json.dump(data, f, indent=2)

    print(f"\nFound {len(data['section_totals'])} section-total entities, "
          f"{len(data.get('sub_allocations', {}))} sub-allocations")
    print(f"Fiscal years: {', '.join(data['fiscal_years'])}")
    for fy in data['fiscal_years']:
        print(f"Total {fy}: ${data['grand_totals'].get(fy, 0):,.0f}")
    if data['warnings']:
        print(f"\nWarnings ({len(data['warnings'])}):")
        for w in data['warnings']:
            print(f"  - {w}")

    print(f"\nGenerating report -> {out_path}")
    generate_excel_report(data, pdf_path, out_path)
    print("Done.")


if __name__ == '__main__':
    main()
